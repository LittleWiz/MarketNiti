import os
import sys
import pandas as pd
from openpyxl import load_workbook

# Add src to sys.path for direct script execution
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from calculate.performance_calculator import prepare_performance_summary

def export_performance_by_category_excel(output_path):
    df = prepare_performance_summary()
    categories = [str(cat)[:31] for cat in df['Category'].unique()]
    numeric_cols = ['Last', '1d', 'Mtd', 'Ytd', '1y', '2y', '5y', '10y', '20y']

    # Step 1: If file exists, load all sheets except those matching categories
    preserved_sheets = {}
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        for sheet in wb.sheetnames:
            if sheet not in categories:
                preserved_sheets[sheet] = pd.DataFrame(wb[sheet].values)
                preserved_sheets[sheet].columns = preserved_sheets[sheet].iloc[0]
                preserved_sheets[sheet] = preserved_sheets[sheet][1:]

    # Step 2: Write all sheets (preserved + new category sheets)
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        # Write preserved sheets first
        for sheet, df_preserved in preserved_sheets.items():
            df_preserved.to_excel(writer, sheet_name=sheet, index=False)

        # Write new category sheets
        for category in categories:
            df_cat = df[df['Category'] == category].copy()
            df_cat = df_cat.rename(columns={'Company': 'Company Names'})
            df_cat = df_cat.drop(columns=['Category'])
            sheet_name = str(category)[:31]
            df_cat.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply number format to numeric columns
            worksheet = writer.sheets[sheet_name]
            workbook = writer.book
            num_format = workbook.add_format({'num_format': '0.0'})

            for idx, col in enumerate(df_cat.columns):
                if col in numeric_cols:     
                    worksheet.set_column(idx, idx, 12, num_format)
                else:
                    worksheet.set_column(idx, idx, 25)  # Wider for company names

    print(f"Excel file with category tabs updated at: {output_path}")

if __name__ == "__main__":
    output_dir = r"D:\Projects\MarketNiti\data\excel"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "Asset_performance.xlsx")
    export_performance_by_category_excel(output_path)